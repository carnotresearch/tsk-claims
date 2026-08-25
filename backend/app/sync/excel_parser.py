"""
Excel parser for HSK Cashless Claims Tracker.

Reads ClaimsMaster, Query_Denial, and Lookups sheets.
All Excel formula cells (LOS, TATs, Outstanding, Ageing, etc.) are recomputed
server-side from raw date/amount columns — we never trust formula-cached values
because TODAY()-based formulas go stale.

Column mapping is by index (0-based), matching the 79-column layout of ClaimsMaster.
"""
from __future__ import annotations

import hashlib
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO

import openpyxl

logger = logging.getLogger(__name__)

# ── Column index constants (ClaimsMaster sheet, 0-based) ─────────────────────
# These are the physical column positions in the Excel sheet.
# If columns are ever added/removed, only update these constants.

C = {
    # Reference & IDs
    "HSK_REF_ID":              0,   # A - formula =IF(I4="","","TANCL-"&TEXT(ROW()-3,"0000"))
    "MONTH_LABEL":             1,   # B - formula =TEXT(O4,"MMM-YY")
    "IHX_REF_ID":              2,   # C
    "HOSPITAL_NAME":           3,   # D
    "LOCATION":                4,   # E
    "ROHINI_ID":               5,   # F
    "UHID":                    6,   # G
    "IP_NUMBER":               7,   # H
    # Patient Details
    "PATIENT_NAME":            8,   # I
    "PATIENT_CONTACT":         9,   # J
    "INSURED_NAME":           10,   # K
    "EMPLOYEE_CODE":          11,   # L
    "CORPORATE_NAME":         12,   # M
    # Admission & Stay
    "DATE_ADMISSION":         13,   # N
    "DATE_DISCHARGE":         14,   # O
    "LOS_DAYS":               15,   # P - formula (recomputed)
    "PROCEDURE_NAME":         16,   # Q
    "DIAGNOSIS":              17,   # R
    # Payor & Policy
    "PAYER_TYPE":             18,   # S
    "TPA_NAME":               19,   # T
    "INSURER_NAME":           20,   # U
    "POLICY_NO":              21,   # V
    "POLICY_TYPE":            22,   # W
    # Preauth
    "PREAUTH_NO":             23,   # X
    "INITIAL_CLAIM_NO":       24,   # Y
    "PREAUTH_REQUEST_DATE":   25,   # Z
    "PREAUTH_APPROVAL_DATE":  26,   # AA
    "PREAUTH_REQUESTED_AMT":  27,   # AB
    "PREAUTH_APPROVED_AMT":   28,   # AC
    "PREAUTH_COPAY":          29,   # AD
    "PREAUTH_STATUS":         30,   # AE
    "PREAUTH_TAT":            31,   # AF - formula (recomputed)
    # Discharge / Final Bill
    "FINAL_BILL_REQUEST_DATE":  32, # AG
    "FINAL_BILL_APPROVAL_DATE": 33, # AH
    "FINAL_CLAIMED_AMT":        34, # AI
    "FINAL_BILL_APPROVED_AMT":  35, # AJ
    "HOSPITAL_DISCOUNT":        36, # AK
    "PATIENT_PAID_AMT":         37, # AL
    "DISCHARGE_STATUS":         38, # AM
    "DISCHARGE_TAT":            39, # AN - formula (recomputed)
    # Claim Submission
    "SUBMISSION_TYPE":         40,  # AO
    "SUBMISSION_DATE":         41,  # AP
    "SUBMISSION_TAT":          42,  # AQ - formula (recomputed)
    "SUBMISSION_STATUS":       43,  # AR
    "COURIER_AGENCY":          44,  # AS
    "COURIER_DESTINATION":     45,  # AT
    "COURIER_DISPATCH_DATE":   46,  # AU
    "COURIER_AWB":             47,  # AV
    "HOSPITAL_INVOICE_NO":     48,  # AW
    # Query & Denial
    "QUERY_RAISED":            49,  # AX
    "QUERY_RAISED_DATE":       50,  # AY
    "QUERY_REASON":            51,  # AZ
    "QUERY_RESPONSE_DATE":     52,  # BA
    "QUERY_RESOLUTION_TAT":    53,  # BB - formula (recomputed)
    "RESUBMISSION_DATE":       54,  # BC
    "DISALLOWED_AMT":          55,  # BD
    "DENIAL_REASON":           56,  # BE
    "APPEAL_FILED":            57,  # BF
    "APPEAL_DATE":             58,  # BG
    # Settlement & Payment
    "SETTLEMENT_DATE":         59,  # BH
    "SETTLED_AMT":             60,  # BI
    "TDS_AMT":                 61,  # BJ
    "DEDUCTION_AMT":           62,  # BK - formula (recomputed)
    "UTR_NO":                  63,  # BL
    "UTR_DATE":                64,  # BM
    "PAYMENT_RECEIVED_DATE":   65,  # BN
    "PAYMENT_RECEIVED_AMT":    66,  # BO
    "PAYMENT_MODE":            67,  # BP
    "HOSPITAL_RECEIPT_NO":     68,  # BQ
    "PAYMENT_TAT":             69,  # BR - formula (recomputed)
    # Outstanding & Ageing
    "OUTSTANDING_AMT":         70,  # BS - formula (recomputed)
    "AGEING_DAYS":             71,  # BT - formula TODAY()-based (recomputed)
    "AGEING_BUCKET":           72,  # BU - formula (recomputed)
    # Status & Notes
    "FINAL_CLAIM_STATUS":      73,  # BV
    "INSURER_COMMENTS":        74,  # BW
    "HOSPITAL_REMARKS":        75,  # BX
    "UPDATED_BY":              76,  # BY
    "LAST_UPDATED_DATE":       77,  # BZ
    # Col 78 (CA) = unnamed variance =(BO+BJ)-AJ, ignored
}

# Lookups sheet: column index → category name
LOOKUP_COLUMNS = {
    0: "payer_type",
    1: "tpa_names",
    2: "insurance_companies",
    3: "insurer_code",
    4: "govt_schemes",
    5: "corporate_sources",
    6: "preauth_status",
    7: "discharge_status",
    8: "submission_status",
    9: "claim_status",
    10: "payment_mode",
    11: "ageing_buckets",
    12: "yes_no",
    13: "policy_type",
    14: "query_reasons",
    15: "disallowed_reasons",
    16: "submission_type",
    17: "users",
}

# Query_Denial sheet column indices (0-based, row 2 = headers)
QD = {
    "HSK_REF_ID":              1,   # B
    "STAGE":                   8,   # I
    "QUERY_RAISED_DATE":       9,   # J
    "QUERY_REASON_CATEGORY":   10,  # K
    "QUERY_REASON_DESC":       11,  # L
    "ACTION_REQUIRED":         12,  # M
    "RESPONSIBLE_PERSON":      13,  # N
    "TARGET_RESPONSE_DATE":    14,  # O
    "RESPONSE_DATE":           15,  # P
    "RESUBMISSION_DATE":       17,  # R
    "DISALLOWED_AMT":          18,  # S
    "DISALLOWED_REASON":       20,  # U
    "APPEAL_FILED":            21,  # V
    "APPEAL_DATE":             22,  # W
    "APPEAL_OUTCOME":          23,  # X
    "FINAL_RECOVERY":          24,  # Y
    "STATUS":                  26,  # AA
    "REMARKS":                 27,  # AB
}


# ── Data containers ───────────────────────────────────────────────────────────

@dataclass
class ParsedHospital:
    name: str
    location: str | None
    rohini_id: str | None


@dataclass
class ParsedClaim:
    hospital_name: str
    hsk_ref_id: str | None
    month_label: str | None
    ihx_ref_id: str | None
    uhid: str | None
    ip_number: str | None
    patient_name: str | None
    patient_contact: str | None
    insured_name: str | None
    employee_code: str | None
    corporate_name: str | None
    date_admission: date | None
    date_discharge: date | None
    los_days: int | None
    procedure_name: str | None
    diagnosis: str | None
    payer_type: str | None
    tpa_name: str | None
    insurer_name: str | None
    policy_no: str | None
    policy_type: str | None
    preauth_no: str | None
    initial_claim_no: str | None
    preauth_request_date: date | None
    preauth_approval_date: date | None
    preauth_requested_amt: Decimal | None
    preauth_approved_amt: Decimal | None
    preauth_copay: Decimal | None
    preauth_status: str | None
    preauth_tat: int | None
    final_bill_request_date: date | None
    final_bill_approval_date: date | None
    final_claimed_amt: Decimal | None
    final_bill_approved_amt: Decimal | None
    hospital_discount: Decimal | None
    patient_paid_amt: Decimal | None
    discharge_status: str | None
    discharge_tat: int | None
    submission_type: str | None
    submission_date: date | None
    submission_tat: int | None
    submission_status: str | None
    courier_agency: str | None
    courier_destination: str | None
    courier_dispatch_date: date | None
    courier_awb: str | None
    hospital_invoice_no: str | None
    query_raised: bool | None
    query_raised_date: date | None
    query_reason: str | None
    query_response_date: date | None
    query_resolution_tat: int | None
    resubmission_date: date | None
    disallowed_amt: Decimal | None
    denial_reason: str | None
    appeal_filed: bool | None
    appeal_date: date | None
    settlement_date: date | None
    settled_amt: Decimal | None
    tds_amt: Decimal | None
    deduction_amt: Decimal | None
    utr_no: str | None
    utr_date: date | None
    payment_received_date: date | None
    payment_received_amt: Decimal | None
    payment_mode: str | None
    hospital_receipt_no: str | None
    payment_tat: int | None
    outstanding_amt: Decimal | None
    ageing_days: int | None
    ageing_bucket: str | None
    final_claim_status: str | None
    insurer_comments: str | None
    hospital_remarks: str | None
    updated_by: str | None
    last_updated_date: date | None
    raw_row_hash: str


@dataclass
class ParsedQueryDenial:
    hsk_ref_id: str | None
    stage: str | None
    query_raised_date: date | None
    query_reason_category: str | None
    query_reason_desc: str | None
    action_required: str | None
    responsible_person: str | None
    target_response_date: date | None
    response_date: date | None
    resolution_tat: int | None
    resubmission_date: date | None
    disallowed_amt: Decimal | None
    disallowed_reason: str | None
    appeal_filed: bool | None
    appeal_date: date | None
    appeal_outcome: str | None
    final_recovery: Decimal | None
    net_loss: Decimal | None
    status: str | None
    remarks: str | None


@dataclass
class ParsedWorkbook:
    hospitals: list[ParsedHospital] = field(default_factory=list)
    claims: list[ParsedClaim] = field(default_factory=list)
    query_denials: list[ParsedQueryDenial] = field(default_factory=list)
    lookups: dict[str, list[str]] = field(default_factory=dict)  # category → [values]
    parse_errors: list[str] = field(default_factory=list)


# ── Type coercion helpers ─────────────────────────────────────────────────────

def _to_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    # Skip Excel formula strings that were not evaluated
    if s.startswith("="):
        return None
    return s or None


def _to_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        s = val.strip()
        if s.startswith("="):
            return None
        try:
            from dateutil.parser import parse as dateparse
            return dateparse(s).date()
        except Exception:
            return None
    return None


def _to_decimal(val: Any) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if s.startswith("=") or s == "" or s == "-":
            return None
        s = s.replace(",", "").replace("₹", "").strip()
        try:
            return Decimal(s)
        except InvalidOperation:
            return None
    try:
        return Decimal(str(val))
    except InvalidOperation:
        return None


def _to_int(val: Any) -> int | None:
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if s.startswith("=") or s == "":
            return None
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _to_bool(val: Any) -> bool | None:
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().upper()
    if s in ("Y", "YES", "TRUE", "1"):
        return True
    if s in ("N", "NO", "FALSE", "0"):
        return False
    return None


def _row_hash(raw_row: tuple) -> str:
    """MD5 of the raw cell values for change detection."""
    content = "|".join(str(v) for v in raw_row)
    return hashlib.md5(content.encode()).hexdigest()


# ── Computed field formulas ───────────────────────────────────────────────────

def _tat(start: date | None, end: date | None) -> int | None:
    if start and end and end >= start:
        return (end - start).days
    return None


def _compute_deduction(
    approved: Decimal | None,
    settled: Decimal | None,
    tds: Decimal | None,
) -> Decimal | None:
    """
    Excel: =MAX(AJ-(BI+BJ),0)
    = max(final_bill_approved - (settled + tds), 0)
    """
    if approved is None or settled is None:
        return None
    tds_val = tds or Decimal(0)
    return max(approved - (settled + tds_val), Decimal(0))


def _compute_outstanding(
    approved: Decimal | None,
    settled: Decimal | None,
    tds: Decimal | None,
    payment_received: Decimal | None,
) -> Decimal | None:
    """
    Excel: =IF(ISNUMBER(BI), MAX((BI+BJ)-(BO+BJ),0), IF(ISNUMBER(AJ), MAX(AJ-(BO+BJ),0), ""))
    Where BI=settled, BJ=tds, BO=payment_received, AJ=approved
    Simplifies to:
      If settled exists: max(settled - payment_received, 0)
      Else if approved exists: max(approved - (payment_received + tds), 0)
    """
    payment = payment_received or Decimal(0)
    tds_val = tds or Decimal(0)
    if settled is not None:
        return max(settled - payment, Decimal(0))
    if approved is not None:
        return max(approved - (payment + tds_val), Decimal(0))
    return None


def _compute_ageing_days(
    outstanding: Decimal | None,
    approval_date: date | None,
) -> int | None:
    """
    Excel: =IF(AND(ISNUMBER(BS),BS>0),TODAY()-AH,"")
    Always recomputed with today's date.
    """
    if outstanding and outstanding > 0 and approval_date:
        return (date.today() - approval_date).days
    return None


def _compute_ageing_bucket(ageing_days: int | None) -> str | None:
    if ageing_days is None:
        return None
    if ageing_days <= 30:
        return "0-30"
    if ageing_days <= 60:
        return "31-60"
    if ageing_days <= 90:
        return "61-90"
    return "90+"


def _compute_month_label(discharge: date | None, admission: date | None) -> str | None:
    """Excel: =TEXT(O4,"MMM-YY") using discharge date."""
    d = discharge or admission
    if d:
        return d.strftime("%b-%y")  # e.g. "Apr-26"
    return None


def _compute_hsk_ref_id(row_index: int, patient_name: str | None) -> str | None:
    """
    Excel: =IF(I4="","","TANCL-"&TEXT(ROW()-3,"0000"))
    row_index is 1-based (first data row = 1).
    Returns None if no patient name.
    """
    if not patient_name:
        return None
    return f"TANCL-{row_index:04d}"


# ── Main parser ───────────────────────────────────────────────────────────────

def parse_workbook(file_stream: BinaryIO) -> ParsedWorkbook:
    """
    Parse the HSK Excel workbook and return structured data.
    Uses data_only=True to read cached formula values where available,
    then recomputes all TODAY()-dependent and financial derived fields.
    """
    result = ParsedWorkbook()

    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True, read_only=True)
    except Exception as exc:
        result.parse_errors.append(f"Failed to open workbook: {exc}")
        return result

    sheet_names = wb.sheetnames

    if "Lookups" in sheet_names:
        _parse_lookups(wb["Lookups"], result)

    if "ClaimsMaster" in sheet_names:
        _parse_claims_master(wb["ClaimsMaster"], result)
    else:
        result.parse_errors.append("ClaimsMaster sheet not found in workbook.")

    if "Query_Denial" in sheet_names:
        _parse_query_denial(wb["Query_Denial"], result)

    wb.close()
    return result


def _parse_lookups(ws, result: ParsedWorkbook) -> None:
    rows = list(ws.iter_rows(values_only=True))
    # Row 0 = description, Row 1 = headers, Row 2+ = values
    if len(rows) < 2:
        return

    for row in rows[2:]:  # skip description and header rows
        for col_idx, category in LOOKUP_COLUMNS.items():
            if col_idx < len(row):
                val = _to_str(row[col_idx])
                if val:
                    if category not in result.lookups:
                        result.lookups[category] = []
                    if val not in result.lookups[category]:
                        result.lookups[category].append(val)


def _parse_claims_master(ws, result: ParsedWorkbook) -> None:
    rows = list(ws.iter_rows(values_only=True))
    # Row 0 = section headers, Row 1 = totals row, Row 2 = column headers, Row 3+ = data
    if len(rows) < 4:
        result.parse_errors.append("ClaimsMaster has no data rows.")
        return

    hospitals_seen: dict[str, ParsedHospital] = {}
    data_row_index = 0  # 1-based counter for data rows only

    for row_num, row in enumerate(rows[3:], start=4):  # 4 = Excel row 4
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        patient_name = _to_str(row[C["PATIENT_NAME"]])
        hospital_name = _to_str(row[C["HOSPITAL_NAME"]])

        # Skip rows without a hospital or patient (structural/summary rows)
        if not hospital_name or not patient_name:
            continue

        data_row_index += 1

        # Track hospitals
        if hospital_name not in hospitals_seen:
            hospitals_seen[hospital_name] = ParsedHospital(
                name=hospital_name,
                location=_to_str(row[C["LOCATION"]]),
                rohini_id=_to_str(row[C["ROHINI_ID"]]),
            )

        # Raw dates and amounts used for computation
        date_admission = _to_date(row[C["DATE_ADMISSION"]])
        date_discharge = _to_date(row[C["DATE_DISCHARGE"]])
        preauth_request_date = _to_date(row[C["PREAUTH_REQUEST_DATE"]])
        preauth_approval_date = _to_date(row[C["PREAUTH_APPROVAL_DATE"]])
        final_bill_request_date = _to_date(row[C["FINAL_BILL_REQUEST_DATE"]])
        final_bill_approval_date = _to_date(row[C["FINAL_BILL_APPROVAL_DATE"]])
        submission_date = _to_date(row[C["SUBMISSION_DATE"]])
        query_raised_date = _to_date(row[C["QUERY_RAISED_DATE"]])
        query_response_date = _to_date(row[C["QUERY_RESPONSE_DATE"]])
        payment_received_date = _to_date(row[C["PAYMENT_RECEIVED_DATE"]])

        final_bill_approved_amt = _to_decimal(row[C["FINAL_BILL_APPROVED_AMT"]])
        settled_amt = _to_decimal(row[C["SETTLED_AMT"]])
        tds_amt = _to_decimal(row[C["TDS_AMT"]])
        payment_received_amt = _to_decimal(row[C["PAYMENT_RECEIVED_AMT"]])

        # Computed fields (always recalculate — never trust Excel cache)
        los_days = _tat(date_admission, date_discharge)
        preauth_tat = _tat(preauth_request_date, preauth_approval_date)
        discharge_tat = _tat(final_bill_request_date, final_bill_approval_date)
        submission_tat = _tat(final_bill_approval_date, submission_date)
        query_resolution_tat = _tat(query_raised_date, query_response_date)
        payment_tat = _tat(submission_date, payment_received_date)
        deduction_amt = _compute_deduction(final_bill_approved_amt, settled_amt, tds_amt)
        outstanding_amt = _compute_outstanding(
            final_bill_approved_amt, settled_amt, tds_amt, payment_received_amt
        )
        ageing_days = _compute_ageing_days(outstanding_amt, final_bill_approval_date)
        ageing_bucket = _compute_ageing_bucket(ageing_days)

        # hsk_ref_id: prefer cached value from Excel, fallback to computed
        hsk_ref_id_raw = _to_str(row[C["HSK_REF_ID"]])
        if hsk_ref_id_raw and not hsk_ref_id_raw.startswith("="):
            hsk_ref_id = hsk_ref_id_raw
        else:
            hsk_ref_id = _compute_hsk_ref_id(data_row_index, patient_name)

        month_label_raw = _to_str(row[C["MONTH_LABEL"]])
        if month_label_raw and not month_label_raw.startswith("="):
            month_label = month_label_raw
        else:
            month_label = _compute_month_label(date_discharge, date_admission)

        raw_hash = _row_hash(row)

        claim = ParsedClaim(
            hospital_name=hospital_name,
            hsk_ref_id=hsk_ref_id,
            month_label=month_label,
            ihx_ref_id=_to_str(row[C["IHX_REF_ID"]]),
            uhid=_to_str(row[C["UHID"]]),
            ip_number=_to_str(row[C["IP_NUMBER"]]),
            patient_name=patient_name,
            patient_contact=_to_str(row[C["PATIENT_CONTACT"]]),
            insured_name=_to_str(row[C["INSURED_NAME"]]),
            employee_code=_to_str(row[C["EMPLOYEE_CODE"]]),
            corporate_name=_to_str(row[C["CORPORATE_NAME"]]),
            date_admission=date_admission,
            date_discharge=date_discharge,
            los_days=los_days,
            procedure_name=_to_str(row[C["PROCEDURE_NAME"]]),
            diagnosis=_to_str(row[C["DIAGNOSIS"]]),
            payer_type=_to_str(row[C["PAYER_TYPE"]]),
            tpa_name=_to_str(row[C["TPA_NAME"]]),
            insurer_name=_to_str(row[C["INSURER_NAME"]]),
            policy_no=_to_str(row[C["POLICY_NO"]]),
            policy_type=_to_str(row[C["POLICY_TYPE"]]),
            preauth_no=_to_str(row[C["PREAUTH_NO"]]),
            initial_claim_no=_to_str(row[C["INITIAL_CLAIM_NO"]]),
            preauth_request_date=preauth_request_date,
            preauth_approval_date=preauth_approval_date,
            preauth_requested_amt=_to_decimal(row[C["PREAUTH_REQUESTED_AMT"]]),
            preauth_approved_amt=_to_decimal(row[C["PREAUTH_APPROVED_AMT"]]),
            preauth_copay=_to_decimal(row[C["PREAUTH_COPAY"]]),
            preauth_status=_to_str(row[C["PREAUTH_STATUS"]]),
            preauth_tat=preauth_tat,
            final_bill_request_date=final_bill_request_date,
            final_bill_approval_date=final_bill_approval_date,
            final_claimed_amt=_to_decimal(row[C["FINAL_CLAIMED_AMT"]]),
            final_bill_approved_amt=final_bill_approved_amt,
            hospital_discount=_to_decimal(row[C["HOSPITAL_DISCOUNT"]]),
            patient_paid_amt=_to_decimal(row[C["PATIENT_PAID_AMT"]]),
            discharge_status=_to_str(row[C["DISCHARGE_STATUS"]]),
            discharge_tat=discharge_tat,
            submission_type=_to_str(row[C["SUBMISSION_TYPE"]]),
            submission_date=submission_date,
            submission_tat=submission_tat,
            submission_status=_to_str(row[C["SUBMISSION_STATUS"]]),
            courier_agency=_to_str(row[C["COURIER_AGENCY"]]),
            courier_destination=_to_str(row[C["COURIER_DESTINATION"]]),
            courier_dispatch_date=_to_date(row[C["COURIER_DISPATCH_DATE"]]),
            courier_awb=_to_str(row[C["COURIER_AWB"]]),
            hospital_invoice_no=_to_str(row[C["HOSPITAL_INVOICE_NO"]]),
            query_raised=_to_bool(row[C["QUERY_RAISED"]]),
            query_raised_date=query_raised_date,
            query_reason=_to_str(row[C["QUERY_REASON"]]),
            query_response_date=query_response_date,
            query_resolution_tat=query_resolution_tat,
            resubmission_date=_to_date(row[C["RESUBMISSION_DATE"]]),
            disallowed_amt=_to_decimal(row[C["DISALLOWED_AMT"]]),
            denial_reason=_to_str(row[C["DENIAL_REASON"]]),
            appeal_filed=_to_bool(row[C["APPEAL_FILED"]]),
            appeal_date=_to_date(row[C["APPEAL_DATE"]]),
            settlement_date=_to_date(row[C["SETTLEMENT_DATE"]]),
            settled_amt=settled_amt,
            tds_amt=tds_amt,
            deduction_amt=deduction_amt,
            utr_no=_to_str(row[C["UTR_NO"]]),
            utr_date=_to_date(row[C["UTR_DATE"]]),
            payment_received_date=payment_received_date,
            payment_received_amt=payment_received_amt,
            payment_mode=_to_str(row[C["PAYMENT_MODE"]]),
            hospital_receipt_no=_to_str(row[C["HOSPITAL_RECEIPT_NO"]]),
            payment_tat=payment_tat,
            outstanding_amt=outstanding_amt,
            ageing_days=ageing_days,
            ageing_bucket=ageing_bucket,
            final_claim_status=_to_str(row[C["FINAL_CLAIM_STATUS"]]),
            insurer_comments=_to_str(row[C["INSURER_COMMENTS"]]),
            hospital_remarks=_to_str(row[C["HOSPITAL_REMARKS"]]),
            updated_by=_to_str(row[C["UPDATED_BY"]]),
            last_updated_date=_to_date(row[C["LAST_UPDATED_DATE"]]),
            raw_row_hash=raw_hash,
        )
        result.claims.append(claim)

    result.hospitals = list(hospitals_seen.values())
    logger.info(
        "ClaimsMaster parsed: %d claims across %d hospitals",
        len(result.claims),
        len(result.hospitals),
    )


def _parse_query_denial(ws, result: ParsedWorkbook) -> None:
    rows = list(ws.iter_rows(values_only=True))
    # Row 0 = description, Row 1 = headers, Row 2+ = data
    if len(rows) < 3:
        return

    for row in rows[2:]:
        if all(v is None for v in row):
            continue
        hsk_ref_id = _to_str(row[QD["HSK_REF_ID"]]) if len(row) > QD["HSK_REF_ID"] else None
        if not hsk_ref_id:
            continue

        # Skip rows where hsk_ref_id looks like a VLOOKUP formula result placeholder
        if hsk_ref_id.startswith("="):
            continue

        query_raised_date = _to_date(row[QD["QUERY_RAISED_DATE"]]) if len(row) > QD["QUERY_RAISED_DATE"] else None
        response_date = _to_date(row[QD["RESPONSE_DATE"]]) if len(row) > QD["RESPONSE_DATE"] else None
        resolution_tat = _tat(query_raised_date, response_date)

        disallowed_amt = _to_decimal(row[QD["DISALLOWED_AMT"]]) if len(row) > QD["DISALLOWED_AMT"] else None
        final_recovery = _to_decimal(row[QD["FINAL_RECOVERY"]]) if len(row) > QD["FINAL_RECOVERY"] else None
        net_loss = disallowed_amt - final_recovery if (disallowed_amt and final_recovery) else disallowed_amt

        qd = ParsedQueryDenial(
            hsk_ref_id=hsk_ref_id,
            stage=_to_str(row[QD["STAGE"]]) if len(row) > QD["STAGE"] else None,
            query_raised_date=query_raised_date,
            query_reason_category=_to_str(row[QD["QUERY_REASON_CATEGORY"]]) if len(row) > QD["QUERY_REASON_CATEGORY"] else None,
            query_reason_desc=_to_str(row[QD["QUERY_REASON_DESC"]]) if len(row) > QD["QUERY_REASON_DESC"] else None,
            action_required=_to_str(row[QD["ACTION_REQUIRED"]]) if len(row) > QD["ACTION_REQUIRED"] else None,
            responsible_person=_to_str(row[QD["RESPONSIBLE_PERSON"]]) if len(row) > QD["RESPONSIBLE_PERSON"] else None,
            target_response_date=_to_date(row[QD["TARGET_RESPONSE_DATE"]]) if len(row) > QD["TARGET_RESPONSE_DATE"] else None,
            response_date=response_date,
            resolution_tat=resolution_tat,
            resubmission_date=_to_date(row[QD["RESUBMISSION_DATE"]]) if len(row) > QD["RESUBMISSION_DATE"] else None,
            disallowed_amt=disallowed_amt,
            disallowed_reason=_to_str(row[QD["DISALLOWED_REASON"]]) if len(row) > QD["DISALLOWED_REASON"] else None,
            appeal_filed=_to_bool(row[QD["APPEAL_FILED"]]) if len(row) > QD["APPEAL_FILED"] else None,
            appeal_date=_to_date(row[QD["APPEAL_DATE"]]) if len(row) > QD["APPEAL_DATE"] else None,
            appeal_outcome=_to_str(row[QD["APPEAL_OUTCOME"]]) if len(row) > QD["APPEAL_OUTCOME"] else None,
            final_recovery=final_recovery,
            net_loss=net_loss,
            status=_to_str(row[QD["STATUS"]]) if len(row) > QD["STATUS"] else None,
            remarks=_to_str(row[QD["REMARKS"]]) if len(row) > QD["REMARKS"] else None,
        )
        result.query_denials.append(qd)

    logger.info("Query_Denial parsed: %d records", len(result.query_denials))
